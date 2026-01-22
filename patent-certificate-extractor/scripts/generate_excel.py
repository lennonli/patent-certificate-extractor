#!/usr/bin/env python3
"""
Excel generation script for patent certificate data

Creates Excel files from extracted patent certificate information.
"""

import sys
from pathlib import Path
from typing import List, Dict, Any
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
except ImportError:
    print("openpyxl not installed. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def normalize_patent_type(patent_type: str) -> str:
    """
    Normalize patent type name to Chinese for consistent sorting

    Args:
        patent_type: Original patent type name (may be in Chinese or English)

    Returns:
        Normalized patent type name in Chinese
    """
    if not patent_type:
        return ''

    type_lower = patent_type.lower()

    # Invention patents
    if any(keyword in type_lower for keyword in ['invention', 'invent', '发明']):
        return '发明专利'

    # Utility model patents
    if any(keyword in type_lower for keyword in ['utility', '实用', '实用新型']):
        return '实用新型专利'

    # Design patents
    if any(keyword in type_lower for keyword in ['design', '外观']):
        return '外观设计专利'

    # Return original if not recognized
    return patent_type


def get_patent_type_priority(patent_type: str) -> int:
    """
    Get priority for patent type sorting

    Priority order:
    1. 发明专利 (highest priority)
    2. 实用新型专利
    3. 外观设计专利 (lowest priority)

    Args:
        patent_type: Normalized patent type name

    Returns:
        Priority number (lower = higher priority)
    """
    type_lower = (patent_type or '').lower()

    if '发明' in type_lower or 'invention' in type_lower:
        return 1
    elif '实用' in type_lower or 'utility' in type_lower:
        return 2
    elif '外观' in type_lower or 'design' in type_lower:
        return 3
    else:
        return 4  # Unknown types


def parse_application_date(date_str: str) -> datetime:
    """
    Parse application date string to datetime object

    Supports multiple date formats:
    - YYYY-MM-DD
    - YYYY/MM/DD
    - DD/MM/YYYY
    - YYYY年MM月DD日
    - MM-DD-YYYY

    Args:
        date_str: Date string to parse

    Returns:
        datetime object (defaults to epoch if parsing fails)
    """
    if not date_str:
        return datetime.min

    date_formats = [
        '%Y-%m-%d',
        '%Y/%m/%d',
        '%d/%m/%Y',
        '%Y年%m月%d日',
        '%m-%d-%Y',
        '%d-%m-%Y'
    ]

    for fmt in date_formats:
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except (ValueError, TypeError):
            continue

    # If all formats fail, return minimum datetime
    return datetime.min


def sort_patent_data(data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Sort patent data by:
    1. Patent holder (alphabetical order)
    2. Patent type priority (Invention > Utility Model > Design)
    3. Application date (descending, most recent first)

    Args:
        data: List of patent dictionaries

    Returns:
        Sorted list of patent dictionaries
    """
    def sort_key(entry):
        holder = entry.get('权利人', '')
        patent_type = entry.get('专利类型', '')
        app_date = entry.get('申请日期', '')

        # Normalize patent type for sorting
        normalized_type = normalize_patent_type(patent_type)

        # Parse date for sorting
        date_obj = parse_application_date(app_date)

        # Sort key: (holder, type_priority, -date_timestamp)
        return (
            holder,  # Alphabetical order
            get_patent_type_priority(normalized_type),  # Type priority (1=Invention, 2=Utility, 3=Design)
            -date_obj.timestamp() if date_obj != datetime.min else 0  # Negative for descending order
        )

    return sorted(data, key=sort_key)


def create_excel(data: List[Dict[str, Any]], output_path: Path) -> bool:
    """
    Create an Excel file with patent certificate data
    
    Args:
        data: List of dictionaries containing patent information
        output_path: Path for the output Excel file
    
    Returns:
        True if successful, False otherwise
    """
    try:
        # Sort data before creating Excel
        sorted_data = sort_patent_data(data)

        wb = Workbook()
        ws = wb.active
        ws.title = "专利证书信息"

        # Define column headers
        headers = ['专利号', '专利名称', '权利人', '专利类型', '发明人', '申请日期']

        # Write headers with styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Define border styles for grouping
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        thick_top_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='medium'),
            bottom=Side(style='thin')
        )

        # Track previous holder for grouping
        previous_holder = None

        # Write data rows
        for row_num, entry in enumerate(sorted_data, 2):
            current_holder = entry.get('权利人', '')

            # Use thick top border for new holder groups
            border_style = thick_top_border if current_holder != previous_holder else thin_border

            ws.cell(row=row_num, column=1, value=entry.get('专利号', '')).border = border_style
            ws.cell(row=row_num, column=2, value=entry.get('专利名称', '')).border = border_style
            ws.cell(row=row_num, column=3, value=entry.get('权利人', '')).border = border_style
            ws.cell(row=row_num, column=4, value=entry.get('专利类型', '')).border = border_style
            ws.cell(row=row_num, column=5, value=entry.get('发明人', '')).border = border_style
            ws.cell(row=row_num, column=6, value=entry.get('申请日期', '')).border = border_style

            previous_holder = current_holder
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0

            # Check header length
            max_length = max(max_length, len(headers[col_num - 1]))

            # Check data length
            for row_num in range(2, len(sorted_data) + 2):
                cell_value = str(ws.cell(row=row_num, column=col_num).value or '')
                max_length = max(max_length, len(cell_value))

            # Set column width (with some padding)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Apply alignment to all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # Save the workbook
        wb.save(output_path)
        return True
        
    except Exception as e:
        print(f"Error creating Excel file: {e}", file=sys.stderr)
        return False


def open_excel(file_path: Path):
    """
    Open the Excel file using the default system application
    
    Args:
        file_path: Path to the Excel file to open
    """
    import subprocess
    import platform
    
    try:
        if platform.system() == 'Darwin':  # macOS
            subprocess.run(['open', str(file_path)], check=True)
        elif platform.system() == 'Windows':
            os.startfile(str(file_path))
        else:  # Linux
            subprocess.run(['xdg-open', str(file_path)], check=True)
    except Exception as e:
        print(f"Could not open Excel file automatically: {e}", file=sys.stderr)
        print(f"Please open the file manually: {file_path}")


def main():
    """Command-line interface for the Excel generation script"""
    if len(sys.argv) < 2:
        print("Usage: python generate_excel.py <output_path>")
        print("Example: python generate_excel.py patent_data.xlsx")
        sys.exit(1)
    
    # For testing with JSON input
    if len(sys.argv) == 3:
        import json
        data = json.loads(sys.argv[2])
    else:
        # Example data for testing
        data = []
    
    output_path = Path(sys.argv[1])
    
    if create_excel(data, output_path):
        print(f"Excel file created successfully: {output_path}")
        open_excel(output_path)
    else:
        print("Failed to create Excel file", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    import os
    main()
